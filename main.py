from bs4 import BeautifulSoup
import pandas as pd

from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from datetime import datetime, timedelta
import xlwings as xw
from openpyxl import load_workbook
import json
import os


wb = load_workbook("Scrape_Control.xlsm")
ws = wb.active
pre_d = [tuple(cell for cell in data if cell is not None)
    for data in ws.iter_rows(values_only=True) if any(cell for cell in data)]
d = dict(zip(pre_d[0],pre_d[1]))
search_word = d["keyword"]
num_item_look_up = int(d["N item to look up"]) // 60
num_co_words = int(d["N of corresponding words"])
print(search_word)
print(num_item_look_up)
print(type(num_item_look_up))
print(num_co_words)
print((type(num_co_words)))
sheet = wb["Scrape_Control"]


def setup_browser():
    options = Options()
    options.binary_location = r"C:\Program Files\Mozilla Firefox\firefox.exe"
    driver_path = r"C:\Users\kayug\Desktop\myvenv1\drivers\geckodriver.exe"
    browser = webdriver.Firefox(service=Service(driver_path), options=options)
    return browser

def open_browser(browser):
    l_by_id = ["gh-la","gh-eb-Geo-a-en",]
    l_by_css = [".gh-eb-Geo-flag.gh-sprRetina",".gh-eb-li-a.gh-icon",".menu-button__button.btn.btn--secondary",
                ".menu-button__item span[data-country='USA|US']",".shipto__close-btn"]
    wait = WebDriverWait(browser, 10)
    url = "https://www.ebay.com/"
    browser.get(url)
    ebay_home_logo = browser.find_element(By.ID, l_by_id[0])
    ebay_home_logo.click()
    language_toggle_1 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, l_by_css[0])))
    language_toggle_1.click()
    language_toggle_2 = wait.until(EC.element_to_be_clickable((By.ID, l_by_id[1])))
    language_toggle_2.click()
    sleep(2)  # for stable functioning
    for element in l_by_css[1:]:
        element_finder_id = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,element)))
        element_finder_id.click()
        sleep(1)

def set_parameter(search_word, browser):
    l_el_id = ["_nkw", "s0-1-17-4[0]-7[1]-_in_kw", "s0-1-17-5[1]-[2]-LH_Sold",
               "s0-1-17-6[3]-[2]-LH_BIN", "s0-1-17-6[4]-[1]-LH_ItemCondition", "s0-1-17-6[7]-[4]-LH_LocatedIn",
               "s0-1-17-6[7]-5[@field[]]-_salic"]
    l_el_css = [".btn.btn--primary"]
    wait = WebDriverWait(browser, 10)
    advanced_setting = wait.until(EC.element_to_be_clickable((By.ID, "gh-as-a")))
    advanced_setting.click()
    for i in range(len(l_el_id)):
        element_finder = browser.find_element(By.ID, l_el_id[i])
        if i == 0:
            element_finder.click()
            element_finder.send_keys(search_word)
        elif i == 1:
            select_keywords_option = Select(element_finder)
            select_keywords_option.select_by_visible_text("Any words, any order")
        elif i == 6:
            select_located_in = Select(element_finder)
            select_located_in.select_by_visible_text("Japan")
        else:
            element_finder.click()

    enter_button = browser.find_element(By.CSS_SELECTOR, l_el_css[0])
    enter_button.click()


def date_extraction(browser):
    html = browser.page_source
    soup = BeautifulSoup(html, "html.parser")
    soup_all = soup.find_all("div", attrs={"class": "s-item__info clearfix"}) # 各商品の文章要素全体
    item_date = None  # item_dateをループ外で使うために
    agg_list = []
    i = 0
    while i < num_item_look_up: #ページをめくる回数
        for j in range(len(soup_all)):  # 商品の名前と日付を取得するループ
            item_x = soup_all[j]
            item_name_0 = item_x.find("span", attrs={"role": "heading"})
            item_date_0 = item_x.find("span", attrs={"class": "POSITIVE"})
            if item_date_0:
                item_date = item_date_0.text.split(" ", 2)[2]
            else:
                print(f"Date not found for item {j}")  # もしデータが日付ではなかった場合の抜け道
            item_date = str(item_date)  # Noneをstrに
            item_name = item_name_0.text
            temp_dict = {}
            temp_dict[item_date] = item_name
            agg_list.append(temp_dict)  # 日付が同じものがオーバーライドされないために

        try:  # 次へボタンのクリック
            next_button = WebDriverWait(browser, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".pagination__next.icon-link"))
            )

            # Scroll to the Next button and click it
            sleep(8) #ここが早すぎるとすぐにブロックされる
            browser.execute_script("arguments[0].scrollIntoView();", next_button)
            next_button.click()
        except Exception as e:
            print("An error occurred while clicking the Next button:", e)
        i += 1
    return agg_list
def data_sorting(agg_list, browser):
    one_month_ago = datetime.now() - timedelta(days=30)
    month_list = [{date: item} for temp_dict in agg_list for date, item in temp_dict.items() if
                  date != "None" and datetime.strptime(date, '%b %d, %Y') > one_month_ago]
    # 同じワードを含むものを抽出
    print(len(month_list))
    all_list = []
    counter= 1
    i = 1
    items = [item_names for temp_dict in month_list for item_names in temp_dict.values()]

    for index_1,item_1 in enumerate(items): #取得してきたデータの商品名部分のみ抽出
        item_1_added = False
        split_item_1 = set(item_1.split())
        for index_2, item_2 in enumerate(items[index_1+1:], start=index_1+1):
            split_item_2 = set(item_2.split())
            common_words = split_item_2.intersection(split_item_1)
            if len(common_words) >= num_co_words:
                if not item_1_added:
                    modified_dict = {k: f"({counter})、{v}" for k, v in month_list[index_1].items()}
                    all_list.append(modified_dict)
                    counter += 1
                    all_list.append(month_list[index_2])
                    item_1_added = True
                else:
                    all_list.append(month_list[index_2])


    print("all_list:",all_list)
    print(len(all_list))
    wb = xw.Book.caller()  # This gets the calling Excel workbook
    ws = wb.sheets['Scrape_Control']  # This gets the Sheet1 of that workbook

    ws.range('D8').value = len(all_list)
    browser.quit()
    return all_list
def sort_for_csv(all_list):
    final_data = []
    for sub_dict in all_list:
        for date, item in sub_dict.items():
            temp_dict = {}
            temp_dict["Date Sold"] = date
            temp_dict["Name of Item"] = item
            final_data.append(temp_dict)
    print(final_data)
    wb = xw.Book("Scrape_Control.xlsm")
    ws = wb.sheets['TempData']
    ws.range('A1').value = json.dumps(final_data)


def main_function():
    browser=setup_browser()
    open_browser(browser)
    set_parameter(search_word,browser)
    agg_list=date_extraction(browser)
    all_list = data_sorting(agg_list,browser)
    sort_for_csv(all_list)

def export_to_csv():
    wb = xw.Book("Scrape_Control.xlsm")
    ws = wb.sheets['TempData']
    stored_value = ws.range('C2').value
    final_data = json.loads(stored_value)
    cwd = os.getcwd()
    try:
        wb = xw.Book("Scrape_Control.xlsm")
        ws = wb.sheets['Scrape_Control']
        file_name = ws.range('D11').value
        if file_name is None:
            raise Exception("File name is None.")
        if not final_data:
            raise Exception("final_data is empty.")
        print(f"the exporting file name is {file_name}")
        df = pd.DataFrame(final_data)
        df.index = df.index + 1
        df.to_csv(fr"INSERT_YOUR_DIRECTORY_HERE{file_name}.csv"", encoding="utf-8-sig", index_label="Index")
    except Exception as e:
        print(f"An error occurred: {e}")





